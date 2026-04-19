#!/usr/bin/env python3
"""
Windows 11 KB Release Notes – Daily Report Generator
Nutzt GitHub Copilot (Models API) statt Anthropic.
Der GITHUB_TOKEN ist in Actions automatisch verfügbar – kein extra Secret nötig.
"""

import os
import re
import json
import smtplib
import requests
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openai import OpenAI
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Konfiguration aus Umgebungsvariablen ───────────────────────────────────
SMTP_SERVER   = os.environ["SMTP_SERVER"]
SMTP_PORT     = int(os.environ.get("SMTP_PORT", 587))
SMTP_USER     = os.environ["SMTP_USER"]
SMTP_PASSWORD = os.environ["SMTP_PASSWORD"]
EMAIL_TO      = os.environ["EMAIL_TO"]          # Kommagetrennt für mehrere Empfänger
EMAIL_FROM    = os.environ.get("EMAIL_FROM", SMTP_USER)
GITHUB_TOKEN  = os.environ["GITHUB_TOKEN"]      # In GitHub Actions automatisch vorhanden!

# GitHub Models API (OpenAI-kompatibler Endpunkt, kostenlos mit GitHub-Account)
GITHUB_MODELS_URL = "https://models.inference.ai.azure.com"
COPILOT_MODEL     = "gpt-4o"                    # Alternativ: "gpt-4o-mini"

HISTORY_URL  = "https://support.microsoft.com/de-de/help/5045988"
HEADERS      = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
LAST_KB_FILE = "last_processed_kb.txt"


# ── Schritt 1: Neueste KB ermitteln ────────────────────────────────────────
def get_latest_kb() -> tuple[str, str, str]:
    resp = requests.get(HISTORY_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    for a in soup.find_all("a", href=True):
        m = re.search(r'/help/(\d{7})', a["href"])
        if m:
            kb_num = "KB" + m.group(1)
            kb_url = "https://support.microsoft.com" + a["href"]
            return kb_num, kb_url, a.get_text(strip=True)
    raise RuntimeError("Keine KB auf der Update-Verlaufsseite gefunden.")

def load_last_kb() -> str:
    if os.path.exists(LAST_KB_FILE):
        with open(LAST_KB_FILE) as f:
            return f.read().strip()
    return ""

def save_last_kb(kb: str):
    with open(LAST_KB_FILE, "w") as f:
        f.write(kb)


# ── Schritt 2: KB-Seite abrufen ───────────────────────────────────────────
def fetch_kb_page(url: str) -> str:
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    main = soup.find("main") or soup.find("article") or soup.body
    return main.get_text(separator="\n", strip=True)[:15000]


# ── Schritt 3: Release Notes via GitHub Copilot extrahieren ───────────────
SYSTEM_PROMPT = """Du bist ein Spezialist im Windows Release Management.
Analysiere den folgenden Microsoft KB-Supportartikel und extrahiere ALLE aufgeführten
Features, Verbesserungen und Bugfixes als strukturierte JSON-Liste.

Antworte NUR mit gültigem JSON – ohne Markdown-Backticks, ohne Präambel, nur das JSON-Objekt.

Erwartetes Format:
{
  "kb": "KB1234567",
  "datum": "TT.MM.JJJJ",
  "builds": "26200.XXXX / 26100.XXXX",
  "eintraege": [
    {
      "rollout": "Normal",
      "titel": "Kurzer Titel (max. 60 Zeichen)",
      "beschreibung": "Vollständige Beschreibung auf Deutsch"
    }
  ]
}

Regeln:
- rollout = "Normal"  → Normaler Rollout (Bugfixes, breite Verfügbarkeit)
- rollout = "Gradual" → Gradueller Rollout (Windows 11 PC-Erfahrungen, neue Features)
- Alle englischen Texte vollständig ins Deutsche übersetzen
- Vollständigkeit: keinen einzigen Eintrag weglassen
- Datum im Format TT.MM.JJJJ"""

def extract_release_notes(kb_text: str, kb_num: str) -> dict:
    client = OpenAI(
        base_url=GITHUB_MODELS_URL,
        api_key=GITHUB_TOKEN,
    )
    response = client.chat.completions.create(
        model=COPILOT_MODEL,
        temperature=0.1,
        max_tokens=4000,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": f"KB-Artikel ({kb_num}):\n\n{kb_text}"}
        ]
    )
    raw = response.choices[0].message.content.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


# ── Schritt 4: Excel generieren ───────────────────────────────────────────
def generate_excel(data: dict) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet0"

    headers = [
        "Einsatz (Monat)", "Einsatz", "Rollout", "KB-Nummer",
        "Betriebssystembuild", "Titel", "Beschreibung",
        "Votum Fachbereich", "Testergebnis", "Bemerkung Client-System", "Aktiv in:"
    ]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    fill_g = PatternFill("solid", fgColor="DEEAF1")  # Blau  = Gradual
    fill_n = PatternFill("solid", fgColor="E2EFDA")  # Grün  = Normal
    dfont  = Font(name="Arial", size=10)

    datum = data.get("datum", "")
    monat = ""
    if datum:
        p = datum.split(".")
        if len(p) == 3:
            monat = f"{p[1].zfill(2)}.{p[2]}"

    for e in data.get("eintraege", []):
        rollout = e.get("rollout", "Gradual")
        ws.append((
            monat, "Microsoft-ClientPatch", rollout,
            data.get("kb",""), data.get("builds",""),
            e.get("titel",""), e.get("beschreibung",""),
            None, None, None, None
        ))
        ri = ws.max_row
        for cell in ws[ri]:
            cell.font = dfont
            cell.fill = fill_g if rollout == "Gradual" else fill_n
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[ri].height = 55

    for i, w in enumerate([12,22,12,14,26,38,90,20,20,40,15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    kb_num   = data.get("kb", "KB_unbekannt")
    filename = f"{kb_num}_Releaseinhalte_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename


# ── Schritt 5: E-Mail senden ──────────────────────────────────────────────
def send_email(filename: str, data: dict, is_new: bool):
    kb         = data.get("kb", "Unbekannt")
    datum      = data.get("datum", datetime.now().strftime("%d.%m.%Y"))
    anzahl     = len(data.get("eintraege", []))
    recipients = [r.strip() for r in EMAIL_TO.split(",")]

    msg = MIMEMultipart()
    msg["From"]    = EMAIL_FROM
    msg["To"]      = ", ".join(recipients)
    msg["Subject"] = f"Windows 11 Release Notes – {kb} ({datum})"

    body = (
        f"Guten Tag,\n\n"
        f"{'🆕 Neues Update verfügbar!' + chr(10) + chr(10) if is_new else ''}"
        f"anbei die automatisch generierten Release Notes für {kb} "
        f"(Windows 11 24H2/25H2, {datum}).\n\n"
        f"Das Update enthält {anzahl} dokumentierte Einträge "
        f"(Gradual Rollout + Normaler Rollout).\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Generiert von GitHub Actions + GitHub Copilot\n"
        f"am {datetime.now().strftime('%d.%m.%Y um %H:%M Uhr')} (UTC)\n"
    )
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(filename, "rb") as f:
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as srv:
        srv.ehlo()
        srv.login(SMTP_USER, SMTP_PASSWORD)
        srv.sendmail(EMAIL_FROM, recipients, msg.as_string())

    print(f"✓ E-Mail mit '{filename}' an {recipients} gesendet.")


# ── Hauptprogramm ─────────────────────────────────────────────────────────
def main():
    force = os.environ.get("FORCE_SEND", "false").lower() == "true"

    print("━━━ Schritt 1: Neueste KB ermitteln ━━━")
    kb, url, title = get_latest_kb()
    print(f"  → {kb}: {title}")
    print(f"  → {url}")

    last_kb = load_last_kb()
    is_new  = kb != last_kb

    if not is_new and not force:
        print(f"\n  ℹ Keine neue KB (letzte verarbeitete: {last_kb}).")
        print("  ℹ Mit FORCE_SEND=true erzwingen.")
        return

    if is_new:
        print(f"  ✓ Neue KB erkannt! (vorherige: {last_kb or 'keine'})")

    print("\n━━━ Schritt 2: KB-Seite abrufen ━━━")
    kb_text = fetch_kb_page(url)
    print(f"  → {len(kb_text)} Zeichen geladen")

    print("\n━━━ Schritt 3: Release Notes extrahieren (GitHub Copilot) ━━━")
    data = extract_release_notes(kb_text, kb)
    print(f"  → {len(data.get('eintraege',[]))} Einträge | Builds: {data.get('builds','?')}")

    print("\n━━━ Schritt 4: Excel generieren ━━━")
    filename = generate_excel(data)
    print(f"  → {filename}")

    print("\n━━━ Schritt 5: E-Mail senden ━━━")
    send_email(filename, data, is_new)

    save_last_kb(kb)
    print(f"\n✅ Fertig! Gespeicherte KB: {kb}")

if __name__ == "__main__":
    main()
